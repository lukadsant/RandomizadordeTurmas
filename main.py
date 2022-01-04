import os
import random
from datetime import datetime
import openpyxl
from openpyxl import Workbook



#Data e Hora
now1 = datetime.now()
dataAtual = now1.strftime("%d-%m-%Y %Hh%M")

#Diretórios

#dirDoExcell = os.path.join(os.path.expandvars("%userprofile%"), "Documents\\Pap appy\\medi.xlsx")
#dirDoExcell2 = os.path.join(os.path.expandvars("%userprofile%"), "Documents\\Pap appy\\rules.xlsx")
#dirExcellDadosSalvos = os.path.join(os.path.expandvars("%userprofile%"),"Documents\\Pap appy\\Dados salvos\\random " + dataAtual + ".xlsx")

dirDoExcell = os.path.join(r"bancox.xlsx")
dirDoExcell2 = os.path.join(r"rulesx.xlsx")
dirExcellDadosSalvos = os.path.join(r"randomizações salvas\randomização " + dataAtual + ".xlsx")

#Openpyxl config
wb = openpyxl.load_workbook(dirDoExcell)
ws = wb.active
book = Workbook()
sheet = book.active
headers = ['Matricula', 'Nome', '','Modulo 1','','Modulo 2','','Modulo 3','','Modulo 4','','Modulo 5']
sheet.append(headers)

wb2 = openpyxl.load_workbook(dirDoExcell2)
#ws2 = wb2['Regra 3 Turmas']
ws3=wb2['Regra ']


#Coletadores de dados
test_list=[]
test_list2=[]
rox=[]
m2=[]
m3=[]
m4=[]
m5=[]



#Função que coleta nome e matricula.
def coletarMatricula():
    for i in range(1, ws.max_row +1):

        test_list.append(ws.cell(i, 2).value)
        test_list2.append(ws.cell(i, 3).value)


def atribuition(qalunosvalue,qlaunos,morex):
    #print(f"tamo na atribuição e o e o qvalue é {qalunosvalue}, esse é o qalunos {qlaunos}")

    divAlunosPorXy=qalunosvalue
    if divAlunosPorXy == 2:
        listOfMorex=[0,0]
        for i in range(morex):
            listOfMorex[i]=+1
        for i in range(qlaunos+listOfMorex[0]): #esse nove é o numero de alunos que cada turma vai ter, é melhor por uma variavel aqui
            rox.append(ws3.cell(59, 1).value)
            m2.append(ws3.cell(59, 2).value)
            m3.append(ws3.cell(59, 3).value)
            m4.append(ws3.cell(59, 4).value)
            m5.append(ws3.cell(59, 5).value)
        for i in range(qlaunos+listOfMorex[1]):
            rox.append(ws3.cell(60, 1).value)
            m2.append(ws3.cell(60, 2).value)
            m3.append(ws3.cell(60, 3).value)
            m4.append(ws3.cell(60, 4).value)
            m5.append(ws3.cell(60, 5).value)
    if divAlunosPorXy == 1:
        listOfMorex=[0]
        for i in range(morex):
            listOfMorex[i]=+1
        for i in range(qlaunos+listOfMorex[0]): #esse nove é o numero de alunos que cada turma vai ter, é melhor por uma variavel aqui
            rox.append(ws3.cell(59, 1).value)
            m2.append(ws3.cell(59, 2).value)
            m3.append(ws3.cell(59, 3).value)
            m4.append(ws3.cell(59, 4).value)
            m5.append(ws3.cell(59, 5).value)
    elif divAlunosPorXy == 3:
        listOfMorex=[0,0,0]
        for i in range(morex):
            listOfMorex[i]=+1
        print("morex: ",morex)
        for i in range(qlaunos+listOfMorex[0]):
            rox.append(ws3.cell(52, 1).value)
            m2.append(ws3.cell(52, 2).value)
            m3.append(ws3.cell(52, 3).value)
            m4.append(ws3.cell(52, 4).value)
            m5.append(ws3.cell(52, 5).value)
        for i in range(qlaunos+listOfMorex[1]):
            rox.append(ws3.cell(53, 1).value)
            m2.append(ws3.cell(53, 2).value)
            m3.append(ws3.cell(53, 3).value)
            m4.append(ws3.cell(53, 4).value)
            m5.append(ws3.cell(53, 5).value)
        for i in range(qlaunos+listOfMorex[2]):
            rox.append(ws3.cell(54, 1).value)
            m2.append(ws3.cell(54, 2).value)
            m3.append(ws3.cell(54, 3).value)
            m4.append(ws3.cell(54, 4).value)
            m5.append(ws3.cell(54, 5).value)

    elif divAlunosPorXy == 4:
        listOfMorex=[0,0,0,0]
        for i in range(morex):
            listOfMorex[i]=+1
        for i in range(qlaunos+listOfMorex[0]):
            rox.append(ws3.cell(44, 1).value)
            m2.append(ws3.cell(44, 2).value)
            m3.append(ws3.cell(44, 3).value)
            m4.append(ws3.cell(44, 4).value)
            m5.append(ws3.cell(44, 5).value)
        for i in range(qlaunos+listOfMorex[1]):
            rox.append(ws3.cell(45, 1).value)
            m2.append(ws3.cell(45, 2).value)
            m3.append(ws3.cell(45, 3).value)
            m4.append(ws3.cell(45, 4).value)
            m5.append(ws3.cell(45, 5).value)
        for i in range(qlaunos+listOfMorex[2]):
            rox.append(ws3.cell(46, 1).value)
            m2.append(ws3.cell(46, 2).value)
            m3.append(ws3.cell(46, 3).value)
            m4.append(ws3.cell(46, 4).value)
            m5.append(ws3.cell(46, 5).value)
        for i in range(qlaunos+listOfMorex[3]):
            rox.append(ws3.cell(47, 1).value)
            m2.append(ws3.cell(47, 2).value)
            m3.append(ws3.cell(47, 3).value)
            m4.append(ws3.cell(47, 4).value)
            m5.append(ws3.cell(47, 5).value)

    elif divAlunosPorXy == 5:
        listOfMorex=[0,0,0,0,0]
        for i in range(morex):
            listOfMorex[i]=+1
        for i in range(qlaunos+listOfMorex[0]):
            rox.append(ws3.cell(35, 1).value)
            m2.append(ws3.cell(35, 2).value)
            m3.append(ws3.cell(35, 3).value)
            m4.append(ws3.cell(35, 4).value)
            m5.append(ws3.cell(35, 5).value)
        for i in range(qlaunos+listOfMorex[1]):
            rox.append(ws3.cell(36, 1).value)
            m2.append(ws3.cell(36, 2).value)
            m3.append(ws3.cell(36, 3).value)
            m4.append(ws3.cell(36, 4).value)
            m5.append(ws3.cell(36, 5).value)
        for i in range(qlaunos+listOfMorex[2]):
            rox.append(ws3.cell(37, 1).value)
            m2.append(ws3.cell(37, 2).value)
            m3.append(ws3.cell(37, 3).value)
            m4.append(ws3.cell(37, 4).value)
            m5.append(ws3.cell(37, 5).value)
        for i in range(qlaunos+listOfMorex[3]):
            rox.append(ws3.cell(38, 1).value)
            m2.append(ws3.cell(38, 2).value)
            m3.append(ws3.cell(38, 3).value)
            m4.append(ws3.cell(38, 4).value)
            m5.append(ws3.cell(38, 5).value)
        for i in range(qlaunos+listOfMorex[4]):
            rox.append(ws3.cell(39, 1).value)
            m2.append(ws3.cell(39, 2).value)
            m3.append(ws3.cell(39, 3).value)
            m4.append(ws3.cell(39, 4).value)
            m5.append(ws3.cell(39, 5).value)

    elif divAlunosPorXy == 6:

        listOfMorex=[0,0,0,0,0,0]
        for i in range(morex):
            listOfMorex[i]=+1

        for i in range(qlaunos+listOfMorex[0]):
            rox.append(ws3.cell(25, 1).value)
            m2.append(ws3.cell(25, 2).value)
            m3.append(ws3.cell(25, 3).value)
            m4.append(ws3.cell(25, 4).value)
            m5.append(ws3.cell(25, 5).value)
        for i in range(qlaunos+listOfMorex[1]):
            rox.append(ws3.cell(26, 1).value)
            m2.append(ws3.cell(26, 2).value)
            m3.append(ws3.cell(26, 3).value)
            m4.append(ws3.cell(26, 4).value)
            m5.append(ws3.cell(26, 5).value)
        for i in range(qlaunos+listOfMorex[2]):
            rox.append(ws3.cell(27, 1).value)
            m2.append(ws3.cell(27, 2).value)
            m3.append(ws3.cell(27, 3).value)
            m4.append(ws3.cell(27, 4).value)
            m5.append(ws3.cell(27, 5).value)
        for i in range(qlaunos+listOfMorex[3]):
            rox.append(ws3.cell(28, 1).value)
            m2.append(ws3.cell(28, 2).value)
            m3.append(ws3.cell(28, 3).value)
            m4.append(ws3.cell(28, 4).value)
            m5.append(ws3.cell(28, 5).value)
        for i in range(qlaunos+listOfMorex[4]):
            rox.append(ws3.cell(29, 1).value)
            m2.append(ws3.cell(29, 2).value)
            m3.append(ws3.cell(29, 3).value)
            m4.append(ws3.cell(29, 4).value)
            m5.append(ws3.cell(29, 5).value)

        for i in range(qlaunos+listOfMorex[5]):
            rox.append(ws3.cell(30, 1).value)
            m2.append(ws3.cell(30, 2).value)
            m3.append(ws3.cell(30, 3).value)
            m4.append(ws3.cell(30, 4).value)
            m5.append(ws3.cell(30, 5).value)

    elif divAlunosPorXy == 7:

        listOfMorex=[0,0,0,0,0,0,0]
        for i in range(morex):
            listOfMorex[i]=+1

        for i in range(qlaunos+listOfMorex[0]):
            rox.append(ws3.cell(14, 1).value)
            m2.append(ws3.cell(14, 2).value)
            m3.append(ws3.cell(14, 3).value)
            m4.append(ws3.cell(14, 4).value)
            m5.append(ws3.cell(14, 5).value)
        for i in range(qlaunos+listOfMorex[1]):
            rox.append(ws3.cell(15, 1).value)
            m2.append(ws3.cell(15, 2).value)
            m3.append(ws3.cell(15, 3).value)
            m4.append(ws3.cell(15, 4).value)
            m5.append(ws3.cell(15, 5).value)
        for i in range(qlaunos+listOfMorex[2]):
            rox.append(ws3.cell(16, 1).value)
            m2.append(ws3.cell(16, 2).value)
            m3.append(ws3.cell(16, 3).value)
            m4.append(ws3.cell(16, 4).value)
            m5.append(ws3.cell(16, 5).value)
        for i in range(qlaunos+listOfMorex[3]):
            rox.append(ws3.cell(17, 1).value)
            m2.append(ws3.cell(17, 2).value)
            m3.append(ws3.cell(17, 3).value)
            m4.append(ws3.cell(17, 4).value)
            m5.append(ws3.cell(17, 5).value)
        for i in range(qlaunos+listOfMorex[4]):
            rox.append(ws3.cell(18, 1).value)
            m2.append(ws3.cell(18, 2).value)
            m3.append(ws3.cell(18, 3).value)
            m4.append(ws3.cell(18, 4).value)
            m5.append(ws3.cell(18, 5).value)

        for i in range(qlaunos+listOfMorex[5]):
            rox.append(ws3.cell(19, 1).value)
            m2.append(ws3.cell(19, 2).value)
            m3.append(ws3.cell(19, 3).value)
            m4.append(ws3.cell(19, 4).value)
            m5.append(ws3.cell(19, 5).value)

        for i in range(qlaunos+listOfMorex[6]):
            rox.append(ws3.cell(20, 1).value)
            m2.append(ws3.cell(20, 2).value)
            m3.append(ws3.cell(20, 3).value)
            m4.append(ws3.cell(20, 4).value)
            m5.append(ws3.cell(20, 5).value)

    elif divAlunosPorXy == 8:

        listOfMorex=[0,0,0,0,0,0,0,0]
        for i in range(morex):
            listOfMorex[i]=+1

        for i in range(qlaunos+listOfMorex[0]):
            rox.append(ws3.cell(3, 1).value)
            m2.append(ws3.cell(3, 2).value)
            m3.append(ws3.cell(3, 3).value)
            m4.append(ws3.cell(3, 4).value)
            m5.append(ws3.cell(3, 5).value)
        for i in range(qlaunos+listOfMorex[1]):
            rox.append(ws3.cell(4, 1).value)
            m2.append(ws3.cell(4, 2).value)
            m3.append(ws3.cell(4, 3).value)
            m4.append(ws3.cell(4, 4).value)
            m5.append(ws3.cell(4, 5).value)
        for i in range(qlaunos+listOfMorex[2]):
            rox.append(ws3.cell(5, 1).value)
            m2.append(ws3.cell(5, 2).value)
            m3.append(ws3.cell(5, 3).value)
            m4.append(ws3.cell(5, 4).value)
            m5.append(ws3.cell(5, 5).value)
        for i in range(qlaunos+listOfMorex[3]):
            rox.append(ws3.cell(6, 1).value)
            m2.append(ws3.cell(6, 2).value)
            m3.append(ws3.cell(6, 3).value)
            m4.append(ws3.cell(6, 4).value)
            m5.append(ws3.cell(6, 5).value)
        for i in range(qlaunos+listOfMorex[4]):
            rox.append(ws3.cell(7, 1).value)
            m2.append(ws3.cell(7, 2).value)
            m3.append(ws3.cell(7, 3).value)
            m4.append(ws3.cell(7, 4).value)
            m5.append(ws3.cell(7, 5).value)

        for i in range(qlaunos+listOfMorex[5]):
            rox.append(ws3.cell(8, 1).value)
            m2.append(ws3.cell(8, 2).value)
            m3.append(ws3.cell(8, 3).value)
            m4.append(ws3.cell(8, 4).value)
            m5.append(ws3.cell(8, 5).value)

        for i in range(qlaunos+listOfMorex[6]):
            rox.append(ws3.cell(9, 1).value)
            m2.append(ws3.cell(9, 2).value)
            m3.append(ws3.cell(9, 3).value)
            m4.append(ws3.cell(9, 4).value)
            m5.append(ws3.cell(9, 5).value)

        for i in range(qlaunos+listOfMorex[7]):
            rox.append(ws3.cell(10, 1).value)
            m2.append(ws3.cell(10, 2).value)
            m3.append(ws3.cell(10, 3).value)
            m4.append(ws3.cell(10, 4).value)
            m5.append(ws3.cell(10, 5).value)

#random.shuffle(test_list2)


#Lugar onde as matriculas serão randomizadas
colab=[]
colabnames=[]
colabnames2=[]
colabnames3=[]
colabnames4=[]
colabnames5=[]
colabnames6=[]

def inserirMatriculasNoArray():
    for i in test_list:
        if i != None:
            colab.append(i)
            colabnames2.append(i)
            colabnames3.append(i)
            colabnames4.append(i)
            colabnames5.append(i)
            colabnames6.append(i)

def inserirNomesNoArray():
    for i in test_list2:
        if i != None:
            colabnames.append(i)

def randomizarMatriculas():
    random.shuffle(colabnames2)





def ralunos():
    listNormal = ["1. maria", "2. joão", "3. enzo", "4. zé", "5. roubaram meu nome"],["1. maria", "2. joão", "3. enzo", "4. zé", "5. roubaram meu nome"],["1. maria", "2. joão", "3. enzo", "4. zé", "5. roubaram meu nome"],["1. maria", "2. joão", "3. enzo", "4. zé", "5. roubaram meu nome"],["1. maria", "2. joão", "3. enzo", "4. zé", "5. roubaram meu nome"]

    print(f"👉 Esta é a lista alunos ordenada: {listNormal[0]}")
    print("👇 Abaixo estarão as lista de alunos randomizadas por modulos")
    for i in range(5):
        random.shuffle(listNormal[i])
    randomList = listNormal
    for i in range(5):
        print(f"Lista modulo {i+1}: {randomList[i]}")

def rturmas():
    listNormal = [11,21,31,41,51,61,71,81],[11,21,31,41,51,61,71,81],[11,21,31,41,51,61,71,81],[11,21,31,41,51,61,71,81],[11,21,31,41,51,61,71,81]

    print(f"👉 Esta é a lista de 8 turmas: {listNormal[0]}")
    print("👇 Abaixo estarão as turmas randomizadas por modulos")
    for i in range(5):
        random.shuffle(listNormal[i])
    randomList = listNormal
    for i in range(5):
        print(f"Lista modulo {i+1}: {randomList[i]}")


def doThis(): #pega randomização pre pronta, porem a lista não é filtrada

    print(f"lista de alunos: {colab}")
    print(f"Há no total {len(colab)} alunos.")
    divAlunosPorX=int(input("Você deseja dividir eles por quantas turmas? "))
    print(f"Cada turma ficará com a quantidade de {int(len(colab)/divAlunosPorX)} alunos, sobrando assim {len(colab)%divAlunosPorX} alunos;")

    for i in range(1, len(colab)+1):

        #rox = (ws2.cell(52,1).value,ws2.cell(53,1).value,ws2.cell(54,1).value)
        #rox=[]
        #rox = (ws2.cell(i, 1).value)

        #print(rox)
        rows = (
            (colab[i-1], colabnames[i],rox[i+1], m2[i+1],m3[i+1],m4[i+1],m5[i+1])

        )
        print(rows)


        sheet.append(rows)
        book.save(dirExcellDadosSalvos)


def denovo():
    divAlunosPorX = int(input("Você deseja dividir eles por quantas turmas? "))
    qalunos = int(len(colab) / divAlunosPorX)
    qalunosresto=int(len(colab) % divAlunosPorX)
    question = input(f"Seguindo a tabela regra, cada turma ficará com a quantidade de {int(len(colab) / divAlunosPorX)} alunos, sobrando assim {len(colab) % divAlunosPorX} alunos, deseja confirmar? (s ou n) ")
    if question=="n":
        question = input(
            f"Seguindo a tabela regra, cada turma ficará com a quantidade de {int(len(colab) / divAlunosPorX)} alunos, sobrando assim {len(colab) % divAlunosPorX} alunos, deseja confirmar? (s ou n) ")

    if len(colab) % divAlunosPorX == 0:
        atribuition(divAlunosPorX, qalunos,qalunosresto)
        for i in range(1, len(colab)+2):
            # rox = (ws2.cell(52,1).value,ws2.cell(53,1).value,ws2.cell(54,1).value)
            # rox2=rox
            # rox = (ws2.cell(i, 1).value)

            # print(rox)

            rows = (
                (colab[i], colabnames[i], " ", colabnames2[i], rox[i], colabnames3[i], m2[i], colabnames4[i], m3[i],
                 colabnames5[i], m4[i], colabnames6[i], m5[i],)

            )
            # print(rows)

            sheet.append(rows)
            book.save(dirExcellDadosSalvos)
        print("A randomização foi concluida")
    else:
        print("Atualmente só é possivel formar turmas quando não há nenhum aluno de fora, tente outra divisão.")
        denovo()


def doThis2():

    print(f"lista de alunos: {colab}")
    print(f"Há no total {len(colab)} alunos.")
    divAlunosPorX = int(input("Você deseja dividir eles por quantas turmas? (Escolha entre 2,3,4,5,6,7,8) "))
    while divAlunosPorX >= 9 or divAlunosPorX <= 0:
        divAlunosPorX = int(input("opa numero incorreto, Você deseja dividir eles por quantas turmas? (Escolha entre 2,3,4,5,6,7,8) "))
    qalunosresto = int(len(colab) % divAlunosPorX)
    qalunos = int(len(colab) / divAlunosPorX)

    question = input(
        f"Seguindo a tabela regra, cada turma ficará com a quantidade de {int(len(colab) / divAlunosPorX)} alunos, sobrando assim {len(colab) % divAlunosPorX} alunos, deseja confirmar? (s ou n) ")

    while question == "n":
        divAlunosPorX = int(input("Você deseja dividir eles por quantas turmas? (Escolha entre 2,3,4,5,6,7,8) "))
        while divAlunosPorX >= 9 or divAlunosPorX <= 0:
            divAlunosPorX = int(
                input("opa numero incorreto, Você deseja dividir eles por quantas turmas? (Escolha entre 2,3,4,5,6,7,8) "))
        qalunosresto = int(len(colab) % divAlunosPorX)
        qalunos = int(len(colab) / divAlunosPorX)
        question = input(
            f"Seguindo a tabela regra, cada turma ficará com a quantidade de {int(len(colab) / divAlunosPorX)} alunos, sobrando assim {len(colab) % divAlunosPorX} alunos, deseja confirmar? (s ou n) ")

    if len(colab)%divAlunosPorX == 0:
        atribuition(divAlunosPorX,qalunos,qalunosresto)
        for i in range(0, len(colab)):

            #rox = (ws2.cell(52,1).value,ws2.cell(53,1).value,ws2.cell(54,1).value)
            #rox2=rox
            #rox = (ws2.cell(i, 1).value)

            #print(rox)

            rows = (
                (colab[i], colabnames[i+1]," ",colabnames2[i],rox[i],colabnames3[i],m2[i],colabnames4[i],m3[i],colabnames5[i],m4[i],colabnames6[i],m5[i],)

            )
            #print(rows)


            sheet.append(rows)
            book.save(dirExcellDadosSalvos)
        print("A randomização foi concluida")
    else:
        atribuition(divAlunosPorX,qalunos,qalunosresto)
        for i in range(0, len(colab)):

            #rox = (ws2.cell(52,1).value,ws2.cell(53,1).value,ws2.cell(54,1).value)
            #rox2=rox
            #rox = (ws2.cell(i, 1).value)

            #print(rox)

            rows = (
                (colab[i], colabnames[i+1]," ",colabnames2[i],rox[i],colabnames2[i],m2[i],colabnames2[i],m3[i],colabnames2[i],m4[i],colabnames2[i],m5[i],)

            )
            #print(rows)


            sheet.append(rows)
            book.save(dirExcellDadosSalvos)
        print("A randomização foi concluida")



coletarMatricula()
inserirMatriculasNoArray()
inserirNomesNoArray()
randomizarMatriculas()
doThis2()